"""
Lee los archivos de insumo de multifuncionales: gestión erestrad, consolidado cajeros cuadrados,
y archivo mensual de arqueos MF (lectura y edición).
"""
from pathlib import Path
from glob import glob
from typing import List, Optional, Set, Tuple, Union
from datetime import datetime, date
import shutil
import re
import pandas as pd
import logging

from src.config.cargador_config import CargadorConfig, PROYECTO_ROOT
from src.insumos.arqueos_mf_calendario import meses_libro_candidatos_fecha_descarga

logger = logging.getLogger(__name__)

# Meses en español para el nombre del archivo ARQUEOS MF
MESES_NOMBRE = (
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
)

# Texto fijo en gestión cuando ARQUEOS MF indica cajero cuadrado (prevalece sobre otras observaciones).
OBSERVACIONES_CUADRADO_EN_ARQUEO = "CUADRADO EN ARQUEO"

# DIARIO sin ARQUEO en gestión: sobrante bajo (parametrizable) y sin calificación previa.
NUEVO_ESTADO_SOBRANTE_DIARIO_CONTABLES = "CONTABILIZACION SOBRANTE CONTABLES"
OBSERVACIONES_SOBRANTE_CUADRE_DIARIO = "SE GRABA SOBRANTE DE CUADRE DIARIO."


class LectorInsumos:
    """Carga los archivos Excel de insumo según la configuración."""

    def __init__(self, config: Optional[CargadorConfig] = None):
        self.config = config or CargadorConfig()
        # Tras preparar_copia_gestion_procesada(), las reglas leen/escriben solo esta ruta (copia *_procesado).
        self._ruta_gestion_escritura: Optional[Path] = None

    @staticmethod
    def _col_cajero_en_df(df: pd.DataFrame) -> Optional[str]:
        for c in df.columns:
            if str(c).strip().lower() in ("cajero", "codigo_cajero", "nit"):
                return c
        return None

    @staticmethod
    def _diario_gestion_fila_tiene_calificacion(df: pd.DataFrame, idx) -> bool:
        """True si la fila ya tiene algún valor operativo de calificación (no sobreescribir)."""
        for col in ("ratificar_grabar_diferencia", "justificacion", "nuevo_estado", "observaciones"):
            if col not in df.columns:
                continue
            v = df.at[idx, col]
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            if str(v).strip():
                return True
        return False

    def _ruta_gestion_original_resuelta(self, fecha: Optional[str] = None) -> Path:
        """
        Resuelve el archivo de gestión fuente (nunca el *_procesado).
        Excluye *_procesado* del glob para no tomar la copia como insumo.
        """
        fecha_usar = fecha or self.config.obtener_fecha_proceso()
        ruta = self._ruta_archivo("gestion_erestrad", fecha_usar)
        if ruta.exists() and "_procesado" not in ruta.stem.lower():
            return ruta
        directorio = self._directorio_insumo("gestion_erestrad")
        patron_fallback = str(directorio / f"gestion/gestion_{fecha_usar}_*.xlsx")
        candidatos = [
            Path(p) for p in glob(patron_fallback) if "_procesado" not in Path(p).stem.lower()
        ]
        candidatos = sorted(candidatos, key=lambda p: p.stat().st_mtime, reverse=True)
        if candidatos:
            ruta_alt = candidatos[0]
            logger.info("Usando archivo de gestión alternativo: %s", ruta_alt)
            return ruta_alt
        if ruta.exists():
            return ruta
        raise FileNotFoundError(
            f"No se encontró el archivo: {ruta}. "
            f"Tampoco archivos con patrón (excl. *_procesado): {patron_fallback}"
        )

    def preparar_copia_gestion_procesada(self, fecha: Optional[str] = None) -> Path:
        """
        Copia la gestión original a un archivo con sufijo _procesado antes de aplicar reglas.
        El original no se modifica; leer_gestion_erestrad y reglas posteriores usan la copia.
        """
        orig = self._ruta_gestion_original_resuelta(fecha)
        proc = orig.with_name(f"{orig.stem}_procesado{orig.suffix}")
        try:
            shutil.copy2(orig, proc)
        except PermissionError:
            # Si _procesado está bloqueado por otro proceso, usar un nombre alterno y continuar.
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            proc = orig.with_name(f"{orig.stem}_procesado_{ts}{orig.suffix}")
            shutil.copy2(orig, proc)
        self._ruta_gestion_escritura = proc.resolve()
        logger.info(
            "Gestión: salida de reglas en %s (original sin cambios: %s).",
            proc.name,
            orig.name,
        )
        return proc

    def _ruta_gestion_para_escritura(self, fecha: Optional[str] = None) -> Path:
        if self._ruta_gestion_escritura is not None:
            return self._ruta_gestion_escritura
        return self._ruta_gestion_original_resuelta(fecha)

    def _df_arqueos_mf_union_meses_descarga(
        self,
        fecha_usar: str,
        hoja_arqueos_mf: Union[str, int] = 0,
    ) -> pd.DataFrame:
        """
        Concatena ARQUEOS MF de los meses candidatos a partir de la fecha de descarga/proceso
        (mes calendario de esa fecha y mes anterior), solo si los archivos existen.
        """
        fd = self._parsear_fecha_dd_mm_yyyy(fecha_usar)
        frames = []
        for mm, aa in meses_libro_candidatos_fecha_descarga(fd):
            ruta = self._ruta_arqueos_mf(mm, aa)
            if ruta.exists():
                frames.append(
                    self.leer_arqueos_mf(mm, aa, hoja=hoja_arqueos_mf, crear_si_falta=False)
                )
        if not frames:
            return pd.DataFrame()
        out = pd.concat(frames, ignore_index=True)
        col_caj = self._buscar_columna_arqueos(out, ("Cajero", "cajero"))
        col_fa = self._buscar_columna_arqueos(out, ("Fecha Arqueo", "Fecha arqueo"))
        if col_caj and col_fa:
            out = out.drop_duplicates(subset=[col_caj, col_fa], keep="last")
        return out

    def _directorio_insumo(self, nombre_insumo: str):
        """Directorio base para un insumo: insumo.directorio si existe, sino directorios.insumos."""
        conf = self.config.cargar()
        insumos = conf.get("insumos", {})
        dir_insumo = insumos.get(nombre_insumo, {}).get("directorio")
        if dir_insumo:
            p = Path(dir_insumo)
            return p if p.is_absolute() else (PROYECTO_ROOT / dir_insumo)
        return self.config.obtener_directorio_insumos()

    def _ruta_archivo(self, nombre_insumo: str, fecha: Optional[str] = None) -> Path:
        """Obtiene la ruta del archivo para un insumo y una fecha DD_MM_YYYY (patron desde config)."""
        conf = self.config.cargar()
        insumos = conf.get("insumos", {})
        if nombre_insumo not in insumos:
            raise KeyError(f"Insumo '{nombre_insumo}' no definido en config")
        if fecha is None:
            fecha = self.config.obtener_fecha_proceso()
        patron = insumos[nombre_insumo].get("patron", "").replace("{fecha}", fecha)
        directorio = self._directorio_insumo(nombre_insumo)
        return directorio / patron

    def _aplicar_regla_diario_sobrantes_extremos(self, df: pd.DataFrame, fecha: Optional[str] = None) -> Tuple[pd.DataFrame, pd.Series]:
        """
        En gestión: para tipo_registro DIARIO con sobrantes altos por magnitud, marcar
        columnas operativas según lineamiento contable.

        Si el mismo cajero tiene fila ARQUEO en el archivo, no se aplica esta regla al DIARIO:
        la calificación del DIARIO debe seguir al ARQUEO (reglas posteriores y sincronización).
        """
        regla = self.config.obtener_regla_diario_sobrantes()
        if not regla.get("activo", False):
            return df, pd.Series(False, index=df.index)
        if "tipo_registro" not in df.columns or "sobrantes" not in df.columns:
            return df, pd.Series(False, index=df.index)

        limite_alto = float(regla.get("limite_sobrante_negativo", 1_000_000_000))
        limite_medio_min = float(regla.get("limite_sobrante_medio_minimo", 50_000_000))
        tipo_norm = df["tipo_registro"].astype(str).str.strip().str.upper()
        tipo_diario = tipo_norm.isin(["DIARIO", "DIADIO"])
        sobrantes_num = pd.to_numeric(df["sobrantes"], errors="coerce").fillna(0)
        sobrantes_abs = sobrantes_num.abs()

        # DIARIO sin fila ARQUEO del mismo cajero: aquí sí aplica la regla por magnitud de sobrantes.
        col_cajero_gestion = None
        for c in df.columns:
            if str(c).strip().lower() in ("cajero", "codigo_cajero", "nit"):
                col_cajero_gestion = c
                break
        cajeros_con_arqueo: Set[int] = set()
        if col_cajero_gestion:
            mask_arqueo = tipo_norm == "ARQUEO"
            for idx in df.index[mask_arqueo]:
                try:
                    cajeros_con_arqueo.add(int(float(df.at[idx, col_cajero_gestion])))
                except (TypeError, ValueError):
                    continue
            caj_num = pd.to_numeric(df[col_cajero_gestion], errors="coerce")
            tipo_diario = tipo_diario & ~caj_num.isin(list(cajeros_con_arqueo))

        mask_alto = tipo_diario & (sobrantes_abs >= limite_alto)
        mask_medio = tipo_diario & (sobrantes_abs >= limite_medio_min) & (sobrantes_abs < limite_alto)
        mask = mask_alto | mask_medio

        if not bool(mask.any()):
            return df, mask

        # Conservamos compatibilidad: algunos insumos traen "ratificar_grabar_diferencia" y otros "grabar".
        if "grabar" in df.columns:
            df["grabar"] = df["grabar"].astype("object")
            df.loc[mask, "grabar"] = "No"
        if "ratificar_grabar_diferencia" in df.columns:
            df["ratificar_grabar_diferencia"] = df["ratificar_grabar_diferencia"].astype("object")
            df.loc[mask, "ratificar_grabar_diferencia"] = "No"
        for col in ("justificacion", "nuevo_estado", "observaciones"):
            if col in df.columns:
                df[col] = df[col].astype("object")
        df.loc[mask, "justificacion"] = "Contable"
        df.loc[mask, "nuevo_estado"] = "ERROR EN TRANSMISION DE CONTADORES"
        df.loc[mask_alto, "observaciones"] = "SALDO NO REAL"
        df.loc[mask_medio, "observaciones"] = "EN ESPERA DE ARQUEO"

        # Tramo medio: si el cajero tiene ARQUEO en ARQUEOS MF, copiar Gestión a Realizar en observaciones.
        if bool(mask_medio.any()) and "cajero" in {str(c).strip().lower() for c in df.columns}:
            try:
                fecha_usar = fecha or self.config.obtener_fecha_proceso()
                df_arq = self._df_arqueos_mf_union_meses_descarga(fecha_usar, hoja_arqueos_mf=0)
                col_cajero_arq = self._buscar_columna_arqueos(df_arq, ("Cajero", "cajero"))
                col_gestion_arq = self._buscar_columna_arqueos(df_arq, ("Gestión a Realizar", "Gestion a Realizar"))
                if col_cajero_arq and col_gestion_arq:
                    mapa_gestion = {}
                    for _, r in df_arq.iterrows():
                        try:
                            caj = int(float(r.get(col_cajero_arq)))
                        except (TypeError, ValueError):
                            continue
                        val = r.get(col_gestion_arq)
                        if val is None:
                            continue
                        txt = str(val).strip()
                        if not txt or txt.lower() == "nan":
                            continue
                        mapa_gestion[caj] = txt

                    col_cajero_gestion = None
                    for c in df.columns:
                        if str(c).strip().lower() in ("cajero", "codigo_cajero"):
                            col_cajero_gestion = c
                            break
                    if col_cajero_gestion:
                        sobrescritas = 0
                        for idx in df[mask_medio].index:
                            try:
                                caj = int(float(df.at[idx, col_cajero_gestion]))
                            except (TypeError, ValueError):
                                continue
                            if caj in mapa_gestion:
                                df.at[idx, "observaciones"] = mapa_gestion[caj]
                                sobrescritas += 1
                        if sobrescritas:
                            logger.info(
                                "Regla DIARIO sobrantes tramo medio: %d fila(s) con observación copiada desde ARQUEOS MF (Gestión a Realizar).",
                                sobrescritas,
                            )
            except Exception as e:
                logger.warning("No se pudo copiar observaciones desde ARQUEOS MF para tramo medio: %s", e)

        logger.info(
            "Regla DIARIO sobrantes extremos aplicada a %d fila(s): tramo alto(abs>=%.0f)=%d, tramo medio(abs>=%.0f y <%.0f)=%d.",
            int(mask.sum()),
            limite_alto,
            int(mask_alto.sum()),
            limite_medio_min,
            limite_alto,
            int(mask_medio.sum()),
        )
        return df, mask

    def _persistir_regla_diario_sobrantes_en_excel(
        self,
        ruta: Path,
        mask: pd.Series,
        df: pd.DataFrame,
        hoja: Optional[str] = None,
    ) -> None:
        """Persiste en el archivo de gestión los campos ajustados por la regla."""
        if mask is None or not bool(mask.any()):
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl no disponible; no se persiste la regla en el archivo de gestión.")
            return

        wb = load_workbook(ruta)
        if hoja is not None and hoja in wb.sheetnames:
            ws = wb[hoja]
        else:
            ws = wb[wb.sheetnames[0]]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers_norm = {str(h or "").strip().lower(): i + 1 for i, h in enumerate(headers)}
        idx_grabar = headers_norm.get("grabar")
        idx_ratificar = headers_norm.get("ratificar_grabar_diferencia")
        idx_just = headers_norm.get("justificacion")
        idx_estado = headers_norm.get("nuevo_estado")
        idx_obs = headers_norm.get("observaciones")
        idx_sobrantes = headers_norm.get("sobrantes")
        idx_faltantes = headers_norm.get("faltantes")

        escritos = 0
        for idx in mask[mask].index:
            excel_row = int(idx) + 2  # +1 por encabezado y +1 por índice 0-based
            if idx_grabar:
                ws.cell(row=excel_row, column=idx_grabar).value = df.at[idx, "grabar"] if "grabar" in df.columns else "No"
            if idx_ratificar:
                ws.cell(row=excel_row, column=idx_ratificar).value = (
                    df.at[idx, "ratificar_grabar_diferencia"] if "ratificar_grabar_diferencia" in df.columns else "No"
                )
            if idx_just:
                ws.cell(row=excel_row, column=idx_just).value = df.at[idx, "justificacion"] if "justificacion" in df.columns else "Contable"
            if idx_estado:
                ws.cell(row=excel_row, column=idx_estado).value = df.at[idx, "nuevo_estado"] if "nuevo_estado" in df.columns else "ERROR EN TRANSMISION DE CONTADORES"
            if idx_obs:
                ws.cell(row=excel_row, column=idx_obs).value = df.at[idx, "observaciones"] if "observaciones" in df.columns else "SALDO NO REAL"
            if idx_sobrantes and "sobrantes" in df.columns:
                ws.cell(row=excel_row, column=idx_sobrantes).value = df.at[idx, "sobrantes"]
            if idx_faltantes and "faltantes" in df.columns:
                ws.cell(row=excel_row, column=idx_faltantes).value = df.at[idx, "faltantes"]

            escritos += 1

        try:
            wb.save(ruta)
            logger.info("Cambios de regla persistidos en archivo gestión (%d fila(s)).", escritos)
        except PermissionError:
            logger.warning(
                "No se pudo guardar el archivo de gestión (%s) por PermissionError (¿abierto en Excel?). Se omiten persistencias.",
                ruta,
            )

    def _persistir_solo_observaciones_gestion(
        self,
        ruta: Path,
        mask: pd.Series,
        df: pd.DataFrame,
        hoja: Optional[str] = None,
    ) -> None:
        """Escribe solo la columna observaciones en el archivo de gestión para las filas indicadas."""
        if mask is None or not bool(mask.any()) or "observaciones" not in df.columns:
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl no disponible; no se persisten observaciones en gestión.")
            return

        wb = load_workbook(ruta)
        if hoja is not None and hoja in wb.sheetnames:
            ws = wb[hoja]
        else:
            ws = wb[wb.sheetnames[0]]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers_norm = {str(h or "").strip().lower(): i + 1 for i, h in enumerate(headers)}
        idx_obs = headers_norm.get("observaciones")
        if not idx_obs:
            logger.warning("No hay columna 'observaciones' en gestión; no se persiste CUADRADO EN ARQUEO.")
            return

        escritos = 0
        for idx in mask[mask].index:
            excel_row = int(idx) + 2
            ws.cell(row=excel_row, column=idx_obs).value = df.at[idx, "observaciones"]
            escritos += 1

        try:
            wb.save(ruta)
            logger.info("Observaciones persistidas en archivo gestión (%d fila(s)).", escritos)
        except PermissionError:
            logger.warning(
                "No se pudo guardar el archivo de gestión (%s) por PermissionError (¿abierto en Excel?).",
                ruta,
            )

    @staticmethod
    def _texto_es_gestion_cuadrado_mf(val: object) -> bool:
        """True si 'Gestión a Realizar' en ARQUEOS MF corresponde al cajero cuadrado estándar."""
        from src.procesamiento.pegar_gestion_a_arqueos_mf import TEXTO_CAJERO_CUADRADO

        t = str(val or "").strip().lower()
        if not t:
            return False
        ref = TEXTO_CAJERO_CUADRADO.strip().lower()
        if t == ref:
            return True
        # Variantes de codificación / tildes
        if "cajero cuadrado" in t and "sucursal" in t:
            return True
        return False

    def _cajeros_cuadrado_mf_en_fecha_descarga(
        self,
        fecha_descarga: date,
        df_arq: pd.DataFrame,
    ) -> Set[int]:
        """Códigos de cajero con fila en ARQUEOS MF: fecha descarga = fecha y gestión = cuadrado."""
        if df_arq is None or df_arq.empty:
            return set()
        col_cajero = self._buscar_columna_arqueos(df_arq, ("Cajero", "cajero"))
        col_fecha = self._buscar_columna_arqueos(df_arq, ("Fecha descarga arqueo", "Fecha Descarga Arqueo"))
        col_gestion = self._buscar_columna_arqueos(df_arq, ("Gestión a Realizar", "Gestion a Realizar"))
        if not (col_cajero and col_fecha and col_gestion):
            return set()

        fechas_col = pd.to_datetime(df_arq[col_fecha], errors="coerce").dt.date
        mask_fecha = fechas_col == fecha_descarga
        cajeros: Set[int] = set()
        for idx in df_arq.loc[mask_fecha].index:
            if not self._texto_es_gestion_cuadrado_mf(df_arq.at[idx, col_gestion]):
                continue
            try:
                cajeros.add(int(float(df_arq.at[idx, col_cajero])))
            except (TypeError, ValueError):
                continue
        return cajeros

    def _aplicar_observaciones_cuadrado_en_df(
        self,
        df_gestion: pd.DataFrame,
        fecha_usar: str,
        hoja_arqueos_mf: Union[str, int] = 0,
    ) -> pd.Series:
        """
        Para cada fila de gestión cuyo cajero está cuadrado en ARQUEOS MF (fecha descarga = proceso):
        ARQUEO y DIARIO quedan con la misma calificación operativa:
        - ratificar_grabar_diferencia = No
        - justificacion = Contable
        - nuevo_estado = ERROR EN TRANSMISION DE CONTADORES
        - observaciones = CUADRADO EN ARQUEO
        """
        fecha_desc = self._parsear_fecha_dd_mm_yyyy(fecha_usar)
        df_arq_union = self._df_arqueos_mf_union_meses_descarga(fecha_usar, hoja_arqueos_mf=hoja_arqueos_mf)
        cajeros = self._cajeros_cuadrado_mf_en_fecha_descarga(fecha_desc, df_arq_union)
        if not cajeros:
            return pd.Series(False, index=df_gestion.index)

        col_cajero_gestion = None
        for c in df_gestion.columns:
            if str(c).strip().lower() in ("cajero", "codigo_cajero", "nit"):
                col_cajero_gestion = c
                break
        if not col_cajero_gestion or "observaciones" not in df_gestion.columns:
            return pd.Series(False, index=df_gestion.index)

        caj_num = pd.to_numeric(df_gestion[col_cajero_gestion], errors="coerce")
        mask = caj_num.isin(list(cajeros))
        if not bool(mask.any()):
            return mask

        for col in ("ratificar_grabar_diferencia", "justificacion", "nuevo_estado", "observaciones", "grabar"):
            if col in df_gestion.columns:
                df_gestion[col] = df_gestion[col].astype("object")
        df_gestion.loc[mask, "observaciones"] = OBSERVACIONES_CUADRADO_EN_ARQUEO
        if "ratificar_grabar_diferencia" in df_gestion.columns:
            df_gestion.loc[mask, "ratificar_grabar_diferencia"] = "No"
        if "justificacion" in df_gestion.columns:
            df_gestion.loc[mask, "justificacion"] = "Contable"
        if "nuevo_estado" in df_gestion.columns:
            df_gestion.loc[mask, "nuevo_estado"] = "ERROR EN TRANSMISION DE CONTADORES"
        if "grabar" in df_gestion.columns:
            df_gestion.loc[mask, "grabar"] = "No"
        return mask

    def aplicar_observaciones_cuadrado_desde_arqueos_mf(
        self,
        fecha: Optional[str] = None,
        hoja_gestion: Optional[str] = None,
        hoja_arqueos_mf: Union[str, int] = 0,
    ) -> int:
        """
        Si en ARQUEOS MF (Fecha descarga arqueo = fecha de proceso) la Gestión a Realizar es
        cajero cuadrado, en gestión (ARQUEO y DIARIO de ese cajero) se fijan operativamente:
        ratificar No, justificación Contable, nuevo_estado ERROR EN TRANSMISION DE CONTADORES,
        observaciones CUADRADO EN ARQUEO.

        Debe ejecutarse al final del flujo para prevalecer sobre otras reglas de observaciones.
        """
        fecha_usar = fecha or self.config.obtener_fecha_proceso()
        try:
            ruta = self._ruta_gestion_para_escritura(fecha_usar)
        except FileNotFoundError:
            logger.warning("No se encontró gestión para %s; no se aplicará CUADRADO EN ARQUEO.", fecha_usar)
            return 0

        if hoja_gestion:
            df_gestion = pd.read_excel(ruta, sheet_name=hoja_gestion, engine="openpyxl")
        else:
            df_gestion = pd.read_excel(ruta, engine="openpyxl")

        mask = self._aplicar_observaciones_cuadrado_en_df(df_gestion, fecha_usar, hoja_arqueos_mf=hoja_arqueos_mf)
        if not bool(mask.any()):
            return 0
        self._persistir_regla_diario_sobrantes_en_excel(
            ruta,
            mask,
            df_gestion,
            hoja=hoja_gestion,
        )
        logger.info(
            "Regla observaciones CUADRADO EN ARQUEO: %d fila(s) en gestión (según ARQUEOS MF fecha descarga).",
            int(mask.sum()),
        )
        return int(mask.sum())

    def sincronizar_observaciones_diario_desde_arqueo(
        self,
        fecha: Optional[str] = None,
        hoja_gestion: Optional[str] = None,
    ) -> int:
        """
        Si un cajero tiene fila ARQUEO y fila DIARIO/DIADIO, copia **observaciones** desde ARQUEO
        hacia DIARIO. No copia ratificar/justificación/nuevo_estado: la regla de grabar 279510020
        define valores distintos en ARQUEO vs DIARIO a propósito.

        Se ejecuta al final del flujo para alinear el texto de calificación cuando el DIARIO
        no debió calificarse por la regla independiente de sobrantes (pareja ARQUEO+DIARIO).
        """
        fecha_usar = fecha or self.config.obtener_fecha_proceso()
        try:
            ruta = self._ruta_gestion_para_escritura(fecha_usar)
        except FileNotFoundError:
            logger.warning("No se encontró gestión para %s; no se sincronizan observaciones DIARIO.", fecha_usar)
            return 0

        if hoja_gestion:
            df_gestion = pd.read_excel(ruta, sheet_name=hoja_gestion, engine="openpyxl")
        else:
            df_gestion = pd.read_excel(ruta, engine="openpyxl")

        if "tipo_registro" not in df_gestion.columns or "observaciones" not in df_gestion.columns:
            return 0

        col_cajero_gestion = None
        for c in df_gestion.columns:
            if str(c).strip().lower() in ("cajero", "codigo_cajero", "nit"):
                col_cajero_gestion = c
                break
        if not col_cajero_gestion:
            return 0

        tipo_norm = df_gestion["tipo_registro"].astype(str).str.strip().str.upper()
        caj_num = pd.to_numeric(df_gestion[col_cajero_gestion], errors="coerce")
        mask_arq = tipo_norm == "ARQUEO"
        mask_di = tipo_norm.isin(["DIARIO", "DIADIO"])
        if not bool(mask_arq.any()) or not bool(mask_di.any()):
            return 0

        mask_mod = pd.Series(False, index=df_gestion.index)
        for caj in caj_num[mask_arq].dropna().unique():
            try:
                icaj = int(float(caj))
            except (TypeError, ValueError):
                continue
            idx_arq = df_gestion.index[(caj_num == icaj) & mask_arq]
            idx_di = df_gestion.index[(caj_num == icaj) & mask_di]
            if len(idx_arq) == 0 or len(idx_di) == 0:
                continue
            src = idx_arq[0]
            val_obs = df_gestion.at[src, "observaciones"]
            for did in idx_di:
                df_gestion.at[did, "observaciones"] = val_obs
                mask_mod.at[did] = True

        if not bool(mask_mod.any()):
            return 0

        df_gestion["observaciones"] = df_gestion["observaciones"].astype("object")
        self._persistir_solo_observaciones_gestion(ruta, mask_mod, df_gestion, hoja=hoja_gestion)
        logger.info(
            "Observaciones DIARIO alineadas desde ARQUEO: %d fila(s).",
            int(mask_mod.sum()),
        )
        return int(mask_mod.sum())

    def aplicar_regla_diario_sobrante_bajo_sin_arqueo(
        self,
        fecha: Optional[str] = None,
        hoja_gestion: Optional[str] = None,
    ) -> int:
        """
        Tras las reglas que acoplan DIARIO con ARQUEO (y sincronización de observaciones):
        filas **solo DIARIO** (sin fila ARQUEO del mismo cajero en el archivo) con
        0 < abs(sobrantes) <= límite (config) reciben la calificación de grabado de sobrante
        de cuadre diario, salvo que la fila **ya** tenga algún valor en ratificar/justificación/
        nuevo_estado/observaciones.
        """
        regla = self.config.obtener_regla_diario_sobrante_bajo_sin_arqueo()
        if not regla.get("activo", False):
            return 0
        limite = float(regla.get("limite_max_abs_sobrante", 3_000_000))

        fecha_usar = fecha or self.config.obtener_fecha_proceso()
        try:
            ruta = self._ruta_gestion_para_escritura(fecha_usar)
        except FileNotFoundError:
            logger.warning(
                "No se encontró gestión para %s; no se aplica regla DIARIO sobrante bajo sin ARQUEO.",
                fecha_usar,
            )
            return 0

        if hoja_gestion:
            df = pd.read_excel(ruta, sheet_name=hoja_gestion, engine="openpyxl")
        else:
            df = pd.read_excel(ruta, engine="openpyxl")

        if "tipo_registro" not in df.columns or "sobrantes" not in df.columns:
            return 0

        col_cajero = self._col_cajero_en_df(df)
        if not col_cajero:
            return 0

        tipo_norm = df["tipo_registro"].astype(str).str.strip().str.upper()
        mask_diario = tipo_norm.isin(["DIARIO", "DIADIO"])

        cajeros_con_arqueo: Set[int] = set()
        for idx in df.index[tipo_norm == "ARQUEO"]:
            try:
                cajeros_con_arqueo.add(int(float(df.at[idx, col_cajero])))
            except (TypeError, ValueError):
                continue

        caj_num = pd.to_numeric(df[col_cajero], errors="coerce")
        mask_sin_par_arqueo = mask_diario & ~caj_num.isin(list(cajeros_con_arqueo))

        sobrantes_num = pd.to_numeric(df["sobrantes"], errors="coerce")
        sobrantes_abs = sobrantes_num.abs()
        mask_monto = (sobrantes_abs > 0) & (sobrantes_abs <= limite)
        mask_candidatas = mask_sin_par_arqueo & mask_monto

        idxs_aplicar = [
            idx for idx in df.index[mask_candidatas] if not self._diario_gestion_fila_tiene_calificacion(df, idx)
        ]
        if not idxs_aplicar:
            return 0

        mask_aplicar = pd.Series(False, index=df.index)
        mask_aplicar.loc[idxs_aplicar] = True

        for col in ("ratificar_grabar_diferencia", "justificacion", "nuevo_estado", "observaciones", "grabar"):
            if col in df.columns:
                df[col] = df[col].astype("object")

        df.loc[mask_aplicar, "ratificar_grabar_diferencia"] = "Si"
        df.loc[mask_aplicar, "justificacion"] = "Contable"
        df.loc[mask_aplicar, "nuevo_estado"] = NUEVO_ESTADO_SOBRANTE_DIARIO_CONTABLES
        df.loc[mask_aplicar, "observaciones"] = OBSERVACIONES_SOBRANTE_CUADRE_DIARIO
        if "grabar" in df.columns:
            df.loc[mask_aplicar, "grabar"] = "Si"

        self._persistir_regla_diario_sobrantes_en_excel(
            ruta,
            mask_aplicar,
            df,
            hoja=hoja_gestion,
        )
        logger.info(
            "Regla DIARIO sobrante bajo sin ARQUEO (abs<= %.0f, excl. ya calificados): %d fila(s).",
            limite,
            int(mask_aplicar.sum()),
        )
        return int(mask_aplicar.sum())

    def _parsear_fecha_dd_mm_yyyy(self, fecha: str) -> date:
        """Convierte DD_MM_YYYY a date."""
        partes = str(fecha).strip().replace("-", "_").split("_")
        if len(partes) != 3:
            raise ValueError(f"Fecha debe ser DD_MM_YYYY: {fecha}")
        d, m, a = int(partes[0]), int(partes[1]), int(partes[2])
        return date(a, m, d)

    def aplicar_regla_arqueo_espera_aclarar_sucursal(
        self,
        fecha: Optional[str] = None,
        hoja_gestion: Optional[str] = None,
        hoja_arqueos_mf: Union[str, int] = 0,
    ) -> int:
        """
        Sincroniza en archivo de gestión (ARQUEO y DIARIO del mismo cajero) cuando
        en ARQUEOS MF la columna "Gestión a Realizar" indica
        "ACLARAR DIFERENCIA Y REPETIR EL ARQUEO" (fecha descarga = proceso).

        Regla (misma calificación en ARQUEO, DIARIO y DIADIO):
        - ratificar_grabar_diferencia = "No"
        - justificacion = "Contable"
        - nuevo_estado = "ERROR EN TRANSMISION DE CONTADORES"
        - observaciones = "EN ESPERA DE ACLARAR SUCURSAL"

        Returns:
            Cantidad de registros de gestión modificados.
        """
        fecha_usar = fecha or self.config.obtener_fecha_proceso()
        fecha_descarga = self._parsear_fecha_dd_mm_yyyy(fecha_usar)

        try:
            ruta = self._ruta_gestion_para_escritura(fecha_usar)
        except FileNotFoundError:
            logger.warning("No se encontró archivo de gestión para %s; no se aplicará la regla ARQUEO.", fecha_usar)
            return 0

        # Leer gestión sin aplicar otras reglas (evitar dobles escrituras)
        if hoja_gestion:
            df_gestion = pd.read_excel(ruta, sheet_name=hoja_gestion, engine="openpyxl")
        else:
            df_gestion = pd.read_excel(ruta, engine="openpyxl")

        if "tipo_registro" not in df_gestion.columns:
            return 0

        tipo_norm = df_gestion["tipo_registro"].astype(str).str.strip().str.upper()
        mask_arqueo_o_diario = tipo_norm.isin(["ARQUEO", "DIARIO", "DIADIO"])

        # ARQUEOS MF: unión de libros candidatos (cruce mes descarga / mes arqueo)
        df_arq = self._df_arqueos_mf_union_meses_descarga(fecha_usar, hoja_arqueos_mf=hoja_arqueos_mf)
        col_cajero_arq = self._buscar_columna_arqueos(df_arq, ("Cajero", "cajero"))
        col_gestion_arq = self._buscar_columna_arqueos(df_arq, ("Gestión a Realizar", "Gestion a Realizar"))
        col_fecha_descarga_arq = self._buscar_columna_arqueos(df_arq, ("Fecha descarga arqueo", "Fecha Descarga Arqueo"))
        col_diferencia_arq = self._buscar_columna_arqueos(df_arq, ("Diferencia", "diferencia"))
        col_traza_arq = self._buscar_columna_arqueos(df_arq, ("paso_a_paso_regla",))
        if not (col_cajero_arq and col_gestion_arq and col_fecha_descarga_arq):
            logger.warning("ARQUEOS MF no tiene columnas necesarias para aplicar la regla ARQUEO.")
            return 0

        if df_arq.empty:
            return 0

        # Filtrar por la fecha de descarga indicada (para no arrastrar reglas de días anteriores)
        fecha_descarga_col = pd.to_datetime(df_arq[col_fecha_descarga_arq], errors="coerce").dt.date
        mask_fecha = fecha_descarga_col == fecha_descarga

        def _txt_match(val: object) -> bool:
            txt = str(val or "").strip().upper()
            if not txt:
                return False
            # Permite pequeñas variaciones (p.ej. DIFERENCIA/DIFERENTIA; con o sin "EL").
            tokens = ("ACLARAR", "REPETIR", "ARQUEO", "DIFER")
            return all(t in txt for t in tokens)

        mask_aclarar = mask_fecha & df_arq[col_gestion_arq].apply(_txt_match)
        if not bool(mask_aclarar.any()):
            return 0

        # Mapa cajero -> true (si hay múltiples, se toma igual)
        cajeros = set()
        for v in df_arq.loc[mask_aclarar, col_cajero_arq].tolist():
            try:
                cajeros.add(int(float(v)))
            except (TypeError, ValueError):
                continue
        if not cajeros:
            return 0

        # Encontrar columna cajero en el archivo de gestión
        col_cajero_gestion = None
        for c in df_gestion.columns:
            if str(c).strip().lower() in ("cajero", "codigo_cajero", "nit"):
                col_cajero_gestion = c
                break
        if not col_cajero_gestion:
            return 0

        # Misma calificación en ARQUEO y DIARIO (y DIADIO) para el cajero con ACLARAR en ARQUEOS MF
        caj_num = pd.to_numeric(df_gestion[col_cajero_gestion], errors="coerce")
        mask_final = mask_arqueo_o_diario & caj_num.isin(list(cajeros))
        if not bool(mask_final.any()):
            return 0

        # Normalizar columnas de salida y aplicar cambios
        for col in ("ratificar_grabar_diferencia", "justificacion", "nuevo_estado", "observaciones"):
            if col in df_gestion.columns:
                df_gestion[col] = df_gestion[col].astype("object")

        for idx in df_gestion[mask_final].index:
            if "ratificar_grabar_diferencia" in df_gestion.columns:
                df_gestion.at[idx, "ratificar_grabar_diferencia"] = "No"
            if "justificacion" in df_gestion.columns:
                df_gestion.at[idx, "justificacion"] = "Contable"
            if "nuevo_estado" in df_gestion.columns:
                df_gestion.at[idx, "nuevo_estado"] = "ERROR EN TRANSMISION DE CONTADORES"
            if "observaciones" in df_gestion.columns:
                df_gestion.at[idx, "observaciones"] = "EN ESPERA DE ACLARAR SUCURSAL"

        # Persistir los cambios
        self._persistir_regla_diario_sobrantes_en_excel(ruta, mask_final, df_gestion, hoja=hoja_gestion)
        logger.info("Regla aclarar/sucursal (ARQUEO+DIARIO) persistida: %d fila(s).", int(mask_final.sum()))
        return int(mask_final.sum())

    def aplicar_regla_grabar_sobrante_desde_arqueos_mf(
        self,
        fecha: Optional[str] = None,
        hoja_gestion: Optional[str] = None,
        hoja_arqueos_mf: Union[str, int] = 0,
    ) -> int:
        """
        Sincroniza en archivo de gestión cuando en ARQUEOS MF el texto de "Gestión a Realizar"
        indica que la diferencia se contabiliza centralizadamente en la cuenta 279510020.

        Para el mismo cajero:
        - ARQUEO:
          ratificar_grabar_diferencia = "Si"
          justificacion = "Fisico"
          nuevo_estado = "CONTABILIZACION SOBRANTE FISICO"
          observaciones = "SE GRABA SOBRANTE DE ARQUEO"
        - DIARIO (uno o varios):
          ratificar_grabar_diferencia = "No"
          justificacion = "Contable"
          nuevo_estado = "ERROR EN TRANSMISION DE CONTADORES"
          observaciones = "SE GRABA SOBRANTE DE ARQUEO"

        Regla:
        - Sobrescribe los campos objetivo aunque ya tengan valores.
        """
        import pandas as pd

        fecha_usar = fecha or self.config.obtener_fecha_proceso()
        fecha_descarga = self._parsear_fecha_dd_mm_yyyy(fecha_usar)

        try:
            ruta = self._ruta_gestion_para_escritura(fecha_usar)
        except FileNotFoundError:
            logger.warning("No se encontró archivo de gestión para %s; no se aplicará regla grabar sobrante.", fecha_usar)
            return 0

        if hoja_gestion:
            df_gestion = pd.read_excel(ruta, sheet_name=hoja_gestion, engine="openpyxl")
        else:
            df_gestion = pd.read_excel(ruta, engine="openpyxl")
        if "tipo_registro" not in df_gestion.columns:
            return 0

        # ARQUEOS MF (unión mes descarga / anterior)
        df_arq = self._df_arqueos_mf_union_meses_descarga(fecha_usar, hoja_arqueos_mf=hoja_arqueos_mf)
        col_cajero_arq = self._buscar_columna_arqueos(df_arq, ("Cajero", "cajero"))
        col_gestion_arq = self._buscar_columna_arqueos(df_arq, ("Gestión a Realizar", "Gestion a Realizar"))
        col_fecha_descarga_arq = self._buscar_columna_arqueos(df_arq, ("Fecha descarga arqueo", "Fecha Descarga Arqueo"))
        col_diferencia_arq = self._buscar_columna_arqueos(df_arq, ("Diferencia", "diferencia"))
        col_traza_arq = self._buscar_columna_arqueos(df_arq, ("paso_a_paso_regla", "paso_a_paso_regla"))
        if not (col_cajero_arq and col_gestion_arq and col_fecha_descarga_arq):
            logger.warning("ARQUEOS MF no tiene columnas necesarias para regla grabar sobrante.")
            return 0

        if df_arq.empty:
            return 0

        fechas_desc = pd.to_datetime(df_arq[col_fecha_descarga_arq], errors="coerce").dt.date
        mask_fecha = fechas_desc == fecha_descarga

        def _parsear_diferencia_desde_traza(txt: object) -> Optional[float]:
            t = str(txt or "")
            m_dif = re.search(r"diferencia_actual=\$(-?[0-9\.]+)", t)
            if not m_dif:
                return None
            try:
                return float(m_dif.group(1).replace(".", ""))
            except (TypeError, ValueError):
                return None

        # Filtrar cajeros ARQUEO MF que realmente quedaron en contabilización de sobrante 279510020.
        # Esta lista es la fuente de verdad para sincronizar la calificación en gestión.
        txt_gestion = df_arq[col_gestion_arq].astype(str).str.strip().str.lower()
        mask_graba_sobrante = txt_gestion.str.contains("cuenta de sobrantes 279510020", na=False)
        mask_candidatos = mask_fecha & mask_graba_sobrante
        if not bool(mask_candidatos.any()):
            return 0

        # Para la fecha dada, usar la diferencia de ARQUEOS MF si existe
        # (desde columna Diferencia o fallback desde paso_a_paso_regla) solo en los cajeros candidatos.
        dif_por_cajero = {}
        cajeros_objetivo = set()
        cols_arq = [col_cajero_arq]
        if col_diferencia_arq:
            cols_arq.append(col_diferencia_arq)
        if col_traza_arq:
            cols_arq.append(col_traza_arq)
        for _, r in df_arq.loc[mask_candidatos, cols_arq].iterrows():
            try:
                caj = int(float(r[col_cajero_arq]))
            except (TypeError, ValueError):
                continue
            cajeros_objetivo.add(caj)
            dif = None
            if col_diferencia_arq and pd.notna(r.get(col_diferencia_arq)):
                try:
                    dif = float(r.get(col_diferencia_arq))
                except (TypeError, ValueError):
                    dif = None
            if dif is None and col_traza_arq:
                dif = _parsear_diferencia_desde_traza(r.get(col_traza_arq))
            if dif is not None:
                dif_por_cajero[caj] = dif

        if not cajeros_objetivo:
            return 0

        col_cajero_gestion = None
        for c in df_gestion.columns:
            if str(c).strip().lower() in ("cajero", "codigo_cajero", "nit"):
                col_cajero_gestion = c
                break
        if not col_cajero_gestion:
            return 0

        tipo_norm = df_gestion["tipo_registro"].astype(str).str.strip().str.upper()
        caj_num = pd.to_numeric(df_gestion[col_cajero_gestion], errors="coerce")
        mask_cajero = caj_num.isin(list(cajeros_objetivo))
        mask_arqueo = mask_cajero & (tipo_norm == "ARQUEO")
        mask_diario = mask_cajero & (tipo_norm.isin(["DIARIO", "DIADIO"]))
        mask_objetivo = mask_arqueo | mask_diario
        if not bool(mask_objetivo.any()):
            return 0

        for col in ("ratificar_grabar_diferencia", "justificacion", "nuevo_estado", "observaciones"):
            if col in df_gestion.columns:
                df_gestion[col] = df_gestion[col].astype("object")

        mask_final_arqueo = mask_arqueo
        mask_final_diario = mask_diario
        mask_final = mask_final_arqueo | mask_final_diario
        if not bool(mask_final.any()):
            return 0

        # Aplicar valores
        if "ratificar_grabar_diferencia" in df_gestion.columns:
            df_gestion.loc[mask_final_arqueo, "ratificar_grabar_diferencia"] = "Si"
            df_gestion.loc[mask_final_diario, "ratificar_grabar_diferencia"] = "No"
        if "justificacion" in df_gestion.columns:
            df_gestion.loc[mask_final_arqueo, "justificacion"] = "Fisico"
            df_gestion.loc[mask_final_diario, "justificacion"] = "Contable"
        if "nuevo_estado" in df_gestion.columns:
            df_gestion.loc[mask_final_arqueo, "nuevo_estado"] = "CONTABILIZACION SOBRANTE FISICO"
            df_gestion.loc[mask_final_diario, "nuevo_estado"] = "ERROR EN TRANSMISION DE CONTADORES"
        if "observaciones" in df_gestion.columns:
            # En DIARIO prima la misma observación que ARQUEO para el cajero.
            df_gestion.loc[mask_final, "observaciones"] = "SE GRABA SOBRANTE DE ARQUEO"

        # Copiar siempre la diferencia final de ARQUEOS MF a gestión en ARQUEO y DIARIO.
        # - diferencia <= 0: sobrantes=dif, faltantes=0
        # - diferencia > 0:  faltantes=dif, sobrantes=0
        if "sobrantes" in df_gestion.columns or "faltantes" in df_gestion.columns:
            for idx in df_gestion[mask_final].index:
                try:
                    caj = int(float(df_gestion.at[idx, col_cajero_gestion]))
                except (TypeError, ValueError):
                    continue
                if caj in dif_por_cajero:
                    dif = float(dif_por_cajero[caj])
                    if "sobrantes" in df_gestion.columns:
                        df_gestion.at[idx, "sobrantes"] = dif if dif <= 0 else 0
                    if "faltantes" in df_gestion.columns:
                        df_gestion.at[idx, "faltantes"] = dif if dif > 0 else 0

        self._persistir_regla_diario_sobrantes_en_excel(
            ruta,
            mask_final,
            df_gestion,
            hoja=hoja_gestion,
        )
        logger.info(
            "Regla grabar sobrante desde ARQUEOS MF persistida: %d fila(s) [ARQUEO=%d, DIARIO=%d].",
            int(mask_final.sum()),
            int(mask_final_arqueo.sum()),
            int(mask_final_diario.sum()),
        )
        return int(mask_final.sum())

    def leer_gestion_erestrad(self, fecha: Optional[str] = None, hoja: Optional[str] = None) -> pd.DataFrame:
        """
        Lee el archivo de gestion para la fecha indicada.

        Comportamiento:
        1) Sin preparar_copia_gestion_procesada(): lee la gestión original (misma resolución que antes).
        2) Tras preparar_copia_gestion_procesada() en main: lee/escribe solo el archivo *_procesado.xlsx
           (el original no se modifica).

        Args:
            fecha: DD_MM_YYYY. Si no se pasa, usa config o fecha actual.
            hoja: Nombre de la hoja. Si None, lee la primera.

        Returns:
            DataFrame con el contenido del archivo.
        """
        fecha_usar = fecha or self.config.obtener_fecha_proceso()
        ruta = self._ruta_gestion_para_escritura(fecha_usar)
        logger.info(f"Leyendo gestión erestrad: {ruta}")
        if hoja:
            df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
        else:
            df = pd.read_excel(ruta, engine="openpyxl")
        df, mask_regla = self._aplicar_regla_diario_sobrantes_extremos(df, fecha=fecha_usar)
        if bool(mask_regla.any()):
            self._persistir_regla_diario_sobrantes_en_excel(
                ruta,
                mask_regla,
                df=df,
                hoja=hoja,
            )
        mask_cuad = self._aplicar_observaciones_cuadrado_en_df(df, fecha_usar, hoja_arqueos_mf=0)
        if bool(mask_cuad.any()):
            self._persistir_regla_diario_sobrantes_en_excel(
                ruta,
                mask_cuad,
                df,
                hoja=hoja,
            )
        logger.info(f"Filas: {len(df)}, columnas: {len(df.columns)}")
        return df

    def leer_consolidado_cajeros_cuadrados(self, fecha: Optional[str] = None, hoja: Optional[str] = None) -> pd.DataFrame:
        """
        Lee el archivo consolidado_cajeros_cuadrados_DD_MM_YYYY.xlsx.

        Args:
            fecha: DD_MM_YYYY. Si no se pasa, usa config o fecha actual.
            hoja: Nombre de la hoja. Si None, lee la primera.

        Returns:
            DataFrame con el contenido del archivo.
        """
        ruta = self._ruta_archivo("consolidado_cajeros_cuadrados", fecha)
        if not ruta.exists():
            raise FileNotFoundError(f"No se encontró el archivo: {ruta}")
        logger.info(f"Leyendo consolidado cajeros cuadrados: {ruta}")
        if hoja:
            df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
        else:
            df = pd.read_excel(ruta, engine="openpyxl")
        logger.info(f"Filas: {len(df)}, columnas: {len(df.columns)}")
        return df

    def leer_todos(self, fecha: Optional[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Lee ambos insumos y devuelve (gestion_erestrad, consolidado_cajeros_cuadrados).

        Args:
            fecha: DD_MM_YYYY. Si None, usa config o fecha actual.

        Returns:
            Tupla (df_gestion, df_consolidado).
        """
        df_gestion = self.leer_gestion_erestrad(fecha=fecha)
        df_consolidado = self.leer_consolidado_cajeros_cuadrados(fecha=fecha)
        return df_gestion, df_consolidado

    # -------------------------------------------------------------------------
    # Archivo mensual: MM- ARQUEOS MF MES AÑO.xlsx (ej: 02- ARQUEOS MF FEBRERO 2026.xlsx)
    # -------------------------------------------------------------------------

    def _ruta_arqueos_mf(self, mes: Optional[int] = None, anio: Optional[int] = None) -> Path:
        """Ruta del archivo de arqueos MF (patron_mes_anio desde config). Si mes/anio no se pasan, usa mes/año actual."""
        if mes is None or anio is None:
            hoy = datetime.now()
            mes = mes if mes is not None else hoy.month
            anio = anio if anio is not None else hoy.year
        if not (1 <= mes <= 12):
            raise ValueError(f"mes debe estar entre 1 y 12, recibido: {mes}")
        conf = self.config.cargar()
        patron = conf.get("insumos", {}).get("arqueos_mf", {}).get(
            "patron_mes_anio", "{mm:02d}- ARQUEOS MF {mes_nombre} {yyyy}.xlsx"
        )
        nombre = (
            patron.replace("{mm:02d}", f"{mes:02d}")
            .replace("{mes_nombre}", MESES_NOMBRE[mes - 1])
            .replace("{yyyy}", str(anio))
        )
        directorio = self._directorio_insumo("arqueos_mf")
        return directorio / nombre

    def _rutas_plantilla_arqueos_mf(self, mes_objetivo: int, anio_objetivo: int) -> List[Path]:
        """
        Archivos existentes para clonar estructura (hojas y columnas), en orden de preferencia:
        mes anterior, luego cualquier ARQUEOS MF en el directorio (más reciente primero).
        """
        directorio = self._directorio_insumo("arqueos_mf")
        ruta_obj = self._ruta_arqueos_mf(mes_objetivo, anio_objetivo)
        ordered: List[Path] = []
        seen: Set[Path] = set()

        def add(p: Path) -> None:
            if not p.exists() or p.resolve() == ruta_obj.resolve():
                return
            rp = p.resolve()
            if rp in seen:
                return
            seen.add(rp)
            ordered.append(p)

        if mes_objetivo > 1:
            add(self._ruta_arqueos_mf(mes_objetivo - 1, anio_objetivo))
        else:
            add(self._ruta_arqueos_mf(12, anio_objetivo - 1))
        try:
            candidatos = sorted(
                [p for p in directorio.glob("*.xlsx") if "procesado" not in p.stem.lower()],
                key=lambda x: x.stat().st_mtime,
                reverse=True,
            )
        except OSError:
            candidatos = []
        for p in candidatos:
            n = p.name.upper()
            if "ARQUEOS" in n and "MF" in n:
                add(p)
        return ordered

    def _crear_arqueos_mf_desde_referencia(self, ruta_destino: Path, ruta_referencia: Path) -> None:
        """Nuevo libro vacío: mismas hojas y columnas que la referencia (solo encabezados, sin filas)."""
        ruta_destino.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelFile(ruta_referencia, engine="openpyxl") as xl:
            nombres = xl.sheet_names
        if not nombres:
            raise ValueError(f"La plantilla no tiene hojas: {ruta_referencia}")
        with pd.ExcelWriter(ruta_destino, engine="openpyxl") as writer:
            for nombre in nombres:
                df_head = pd.read_excel(ruta_referencia, sheet_name=nombre, engine="openpyxl", nrows=0)
                vacio = pd.DataFrame(columns=df_head.columns)
                vacio.to_excel(writer, sheet_name=nombre, index=False)

    def asegurar_archivo_arqueos_mf(self, mes: int, anio: int) -> bool:
        """
        Si el ARQUEOS MF del mes/año no existe, crea un Excel vacío en el mismo directorio
        con las mismas hojas y columnas que un archivo ARQUEOS MF de referencia.

        Returns:
            True si se creó el archivo, False si ya existía.

        Raises:
            FileNotFoundError si no hay ningún ARQUEOS MF en el directorio para usar como plantilla.
        """
        if mes is None or anio is None:
            raise ValueError("asegurar_archivo_arqueos_mf requiere mes y anio explícitos")
        ruta = self._ruta_arqueos_mf(mes=mes, anio=anio)
        if ruta.exists():
            return False
        for ref in self._rutas_plantilla_arqueos_mf(mes, anio):
            try:
                self._crear_arqueos_mf_desde_referencia(ruta, ref)
                logger.info(
                    "Creado insumo ARQUEOS MF vacío: %s (columnas según plantilla %s).",
                    ruta.name,
                    ref.name,
                )
                return True
            except Exception as e:
                logger.warning("No se pudo generar ARQUEOS MF desde plantilla %s: %s", ref.name, e)
                continue
        raise FileNotFoundError(
            f"No existe {ruta.name} y no hay un ARQUEOS MF válido en "
            f"{ruta.parent} para copiar la estructura de columnas."
        )

    def leer_arqueos_mf(
        self,
        mes: Optional[int] = None,
        anio: Optional[int] = None,
        hoja: Optional[Union[str, int]] = None,
        crear_si_falta: bool = True,
    ) -> pd.DataFrame:
        """
        Lee el archivo mensual de arqueos MF (ej: 02- ARQUEOS MF FEBRERO 2026.xlsx).

        Args:
            mes: Número de mes 1-12. Si None, mes actual.
            anio: Año (ej: 2026). Si None, año actual.
            hoja: Nombre o índice de hoja. Si None, primera hoja.
            crear_si_falta: Si True y el archivo no existe, crea uno vacío con las mismas hojas/columnas
                que otro ARQUEOS MF del directorio.

        Returns:
            DataFrame con el contenido de la hoja.
        """
        if mes is None or anio is None:
            hoy = datetime.now()
            mes = mes if mes is not None else hoy.month
            anio = anio if anio is not None else hoy.year
        ruta = self._ruta_arqueos_mf(mes=mes, anio=anio)
        if not ruta.exists():
            if crear_si_falta:
                try:
                    self.asegurar_archivo_arqueos_mf(mes, anio)
                except FileNotFoundError:
                    raise FileNotFoundError(f"No se encontró el archivo: {ruta}") from None
            else:
                raise FileNotFoundError(f"No se encontró el archivo: {ruta}")
        logger.info("Leyendo arqueos MF: %s", ruta)
        if hoja is not None:
            df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
        else:
            df = pd.read_excel(ruta, engine="openpyxl")
        logger.info("Filas: %d, columnas: %d", len(df), len(df.columns))
        return df

    def _ruta_historico_cuadre(self) -> Path:
        """Ruta del archivo de historico cuadre (patron desde config)."""
        conf = self.config.cargar()
        insumos = conf.get("insumos", {})
        patron = insumos.get("historico_cuadre_cajeros_sucursales", {}).get("patron", "HISTORICO_CUADRE_CAJEROS_SUCURSALES.xlsx")
        directorio = self._directorio_insumo("historico_cuadre_cajeros_sucursales")
        return directorio / patron

    def leer_historico_cuadre_cajeros_sucursales(self, hoja: Optional[Union[str, int]] = None) -> pd.DataFrame:
        """
        Lee HISTORICO_CUADRE_CAJEROS_SUCURSALES.xlsx (tipo_registro, fecha_arqueo, cajero, etc.).
        Se usa para obtener la fecha del ultimo arqueo cuando en ARQUEOS MF solo hay un registro del cajero.
        """
        ruta = self._ruta_historico_cuadre()
        if not ruta.exists():
            raise FileNotFoundError(f"No se encontro el archivo: {ruta}")
        logger.info("Leyendo historico cuadre cajeros sucursales: %s", ruta)
        if hoja is not None:
            df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
        else:
            df = pd.read_excel(ruta, engine="openpyxl")
        logger.info("Historico: %d filas, %d columnas", len(df), len(df.columns))
        return df

    def guardar_arqueos_mf(
        self,
        df: pd.DataFrame,
        mes: Optional[int] = None,
        anio: Optional[int] = None,
        hoja: Union[str, int] = 0,
    ) -> Path:
        """
        Guarda (edita) el archivo mensual de arqueos MF. Si el archivo tiene varias hojas,
        se reemplaza solo la hoja indicada y se mantienen el resto.

        Args:
            df: DataFrame a escribir.
            mes: Número de mes 1-12. Si None, mes actual.
            anio: Año. Si None, año actual.
            hoja: Nombre o índice de la hoja a reemplazar (default 0).

        Returns:
            Path del archivo guardado.
        """
        ruta = self._ruta_arqueos_mf(mes=mes, anio=anio)
        if not ruta.exists():
            hoy = datetime.now()
            mm = mes if mes is not None else hoy.month
            yy = anio if anio is not None else hoy.year
            try:
                self.asegurar_archivo_arqueos_mf(mm, yy)
            except FileNotFoundError:
                pass
        if ruta.exists():
            # Leer todas las hojas, reemplazar la indicada y volver a escribir
            with pd.ExcelFile(ruta, engine="openpyxl") as xl:
                nombres_hojas = xl.sheet_names
            reemplazar = (lambda i, n: i == hoja) if isinstance(hoja, int) else (lambda i, n: n == hoja)
            dict_hojas = {}
            for i, nombre in enumerate(nombres_hojas):
                if reemplazar(i, nombre):
                    dict_hojas[nombre] = df
                else:
                    dict_hojas[nombre] = pd.read_excel(ruta, sheet_name=nombre, engine="openpyxl")
            with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
                for nombre, frame in dict_hojas.items():
                    frame.to_excel(writer, sheet_name=nombre, index=False)
        else:
            df.to_excel(ruta, index=False, engine="openpyxl")
        logger.info("Guardado arqueos MF: %s", ruta)
        return ruta

    def _buscar_columna_arqueos(self, df: pd.DataFrame, nombres: tuple) -> Optional[str]:
        """Retorna la primera columna cuyo nombre normalizado está en nombres."""
        for c in df.columns:
            if (str(c).strip().lower() in [n.lower() for n in nombres]):
                return c
        return None

    def quitar_filas_por_fecha_descarga_arqueo(
        self,
        fecha_filtro: date,
        mes: Optional[int] = None,
        anio: Optional[int] = None,
        hoja: Union[str, int] = 0,
    ) -> int:
        """
        Borra del ARQUEOS MF todas las filas donde "Fecha descarga arqueo" = fecha_filtro.
        Revisa los libros mensuales candidatos (mes de la fecha y mes anterior). mes/anio se ignoran.

        Returns:
            Número total de filas eliminadas.
        """
        total = 0
        for mm, aa in meses_libro_candidatos_fecha_descarga(fecha_filtro):
            ruta = self._ruta_arqueos_mf(mm, aa)
            if not ruta.exists():
                continue
            df = self.leer_arqueos_mf(mes=mm, anio=aa, hoja=hoja)
            col = self._buscar_columna_arqueos(df, ("Fecha descarga arqueo", "Fecha Descarga Arqueo"))
            if not col:
                logger.warning("No se encontró columna 'Fecha descarga arqueo'; no se puede hacer reset en %s.", ruta.name)
                continue

            def valor_a_fecha(val):
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    return None
                if isinstance(val, date) and not isinstance(val, datetime):
                    return val
                if isinstance(val, datetime):
                    return val.date()
                try:
                    dt = pd.to_datetime(val)
                    return dt.date() if hasattr(dt, "date") else dt
                except Exception:
                    return None

            mask = df[col].apply(lambda v: valor_a_fecha(v) == fecha_filtro)
            n_quitar = int(mask.sum())
            if n_quitar == 0:
                continue
            df_nuevo = df[~mask].copy()
            self.guardar_arqueos_mf(df_nuevo, mes=mm, anio=aa, hoja=hoja)
            logger.info(
                "Reset ARQUEOS MF (%s): eliminadas %d fila(s) con Fecha descarga arqueo = %s.",
                ruta.name,
                n_quitar,
                fecha_filtro,
            )
            total += n_quitar
        if total == 0:
            logger.info("Reset ARQUEOS MF: no hay filas con Fecha descarga arqueo = %s en libros candidatos.", fecha_filtro)
        return total
