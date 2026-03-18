"""
Pega en el archivo ARQUEOS MF registros de:
- archivo de gestión (solo tipo_registro = ARQUEO)
- consolidado_cajeros_cuadrados (solo tipo_cajero que contenga "multifuncional" y tipo_registro = "ARQUEO")
Misma asociación de columnas origen -> ARQUEOS MF.
"""
from pathlib import Path
from typing import Optional, Dict, Any, Union
import pandas as pd
import logging

logger = logging.getLogger(__name__)

# Mapeo definitivo: nombre en origen (gestión o consolidado) -> nombre de columna en ARQUEOS MF
MAPEO_A_ARQUEOS_MF = {
    "codigo_suc": "Sucursal",
    "codigo_cajero": "Cajero",
    "marca": "Marca herramienta",
    "fecha_asignacion": "Fecha descarga arqueo",
    "fecha_arqueo": "Fecha Arqueo",
    "hora_arqueo": "Hora Arqueo",
    "arqueo_fisico/saldo_contadores": "Efectivo Arqueado /Arqueo fisico saldo contadores",
    "saldo_contable": "Saldo Contable",
    "dispensado_corte_arqueo": "dispensado_corte_arqueo",
    "recibido_corte_arqueo": "recibido_corte_arqueo",
    "documento_responsable": "Documento responsable",
}
# Alias en archivos origen (p. ej. consolidado): nombres alternativos que también mapean a la columna ARQUEOS MF
ALIASES_ORIGEN = {
    "codigo_cajero": ["Cajero", "cajero", "NIT", "nit"],
    "fecha_arqueo": ["Fecha Arqueo", "Fecha arqueo", "fecha arqueo"],
    "saldo_contable": ["Saldo contable", "Saldo Contable", "saldo contable"],
    "codigo_suc": ["Sucursal", "sucursal"],
    "marca": ["Marca herramienta", "Marca"],
    "fecha_asignacion": ["Fecha descarga arqueo", "fecha descarga arqueo"],
}


def _normalizar_nombre_columna(s: str) -> str:
    return (s or "").strip().lower()


def _encontrar_columna(df: pd.DataFrame, nombres_posibles: list) -> Optional[str]:
    """Devuelve la primera columna de df cuyo nombre normalizado está en nombres_posibles."""
    norm = [_normalizar_nombre_columna(x) for x in nombres_posibles]
    for c in df.columns:
        if _normalizar_nombre_columna(str(c)) in norm:
            return c
    return None


def _mapear_fila_a_arqueos_mf(row: pd.Series, columnas_origen: dict) -> Dict[str, Any]:
    """Convierte una fila de origen (gestión o consolidado) al formato de fila para ARQUEOS MF."""
    out = {}
    for col_origen, col_arqueos in columnas_origen.items():
        if col_origen in row.index:
            out[col_arqueos] = row[col_origen]
    return out


TEXTO_CAJERO_CUADRADO = "Cajero cuadrado, sin observación por parte de la sucursal"


def _aplicar_texto_cajero_cuadrado(df: pd.DataFrame, tolerancia: float = 0.01) -> int:
    """
    Para cada fila donde Diferencia = Saldo - (Efectivo + dispensado - recibido) - Remanente ≈ 0,
    escribe en Gestión a Realizar: "Cajero cuadrado, sin observación por parte de la sucursal".
    Modifica df in place. Retorna cantidad de filas actualizadas.
    """
    col_saldo = _encontrar_columna(df, ["Saldo Contable", "Saldo contable"])
    col_efectivo = _encontrar_columna(df, ["Efectivo Arqueado /Arqueo fisico saldo contadores"])
    col_dispensado = _encontrar_columna(df, ["dispensado_corte_arqueo"])
    col_recibido = _encontrar_columna(df, ["recibido_corte_arqueo"])
    col_remanente = _encontrar_columna(df, ["Remanente /Provisión /Ajustes", "Remanente/Provisión/Ajustes"])
    col_gestion = _encontrar_columna(df, ["Gestión a Realizar", "Gestion a Realizar"])
    if not all([col_saldo, col_efectivo, col_dispensado, col_recibido]):
        return 0
    if col_gestion is None:
        col_gestion = "Gestión a Realizar"
        if col_gestion not in df.columns:
            df[col_gestion] = ""
    def _float(x, default=0.0):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return default
        try:
            return float(x)
        except (ValueError, TypeError):
            return default
    n = 0
    for idx in df.index:
        saldo = _float(df.at[idx, col_saldo])
        efectivo = _float(df.at[idx, col_efectivo])
        dispensado = _float(df.at[idx, col_dispensado])
        recibido = _float(df.at[idx, col_recibido])
        remanente = _float(df.at[idx, col_remanente]) if col_remanente and col_remanente in df.columns else 0.0
        diferencia = saldo - (efectivo + dispensado - recibido) - remanente
        if abs(diferencia) <= tolerancia:
            # No sobrescribir si ya tiene el texto de cruce faltante-sobrante
            valor_actual = df.at[idx, col_gestion]
            if isinstance(valor_actual, str) and "es cruzado con sobrante" in valor_actual:
                continue
            df.at[idx, col_gestion] = TEXTO_CAJERO_CUADRADO
            n += 1
    return n


def _escribir_formula_diferencia_en_excel(ruta: Path, nombre_hoja: str = "DETALLE MF") -> None:
    """
    Escribe en la columna Diferencia la fórmula:
    =Saldo Contable - (Efectivo Arqueado/Arqueo fisico saldo contadores + dispensado_corte_arqueo - recibido_corte_arqueo) - Remanente/Provisión/Ajustes
    para todas las filas de datos en la hoja.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl no disponible para escribir fórmula Diferencia")
        return
    wb = load_workbook(ruta)
    if nombre_hoja not in wb.sheetnames:
        nombre_hoja = wb.sheetnames[0]
    sheet = wb[nombre_hoja]
    # Fila 1 = encabezados
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    def col_index(nombre: str) -> Optional[int]:
        n = (nombre or "").strip().lower()
        for i, h in enumerate(headers):
            if (str(h or "").strip().lower() == n):
                return i + 1
        return None

    idx_saldo = col_index("Saldo Contable")
    idx_efectivo = col_index("Efectivo Arqueado /Arqueo fisico saldo contadores")
    idx_dispensado = col_index("dispensado_corte_arqueo")
    idx_recibido = col_index("recibido_corte_arqueo")
    idx_remanente = col_index("Remanente /Provisión /Ajustes")
    idx_diferencia = col_index("Diferencia")

    if not all([idx_saldo, idx_efectivo, idx_dispensado, idx_recibido, idx_diferencia]):
        logger.warning("No se encontraron todas las columnas para la fórmula Diferencia; se omite.")
        return
    # idx_remanente puede no existir; en la fórmula usamos 0 si no está
    letra_saldo = get_column_letter(idx_saldo)
    letra_efectivo = get_column_letter(idx_efectivo)
    letra_dispensado = get_column_letter(idx_dispensado)
    letra_recibido = get_column_letter(idx_recibido)
    letra_remanente = get_column_letter(idx_remanente) if idx_remanente else None
    letra_diferencia = get_column_letter(idx_diferencia)

    idx_naturaleza = col_index("Naturaleza")
    # Asegurar no sobrescribir Gestión a Realizar (texto de cruce / cajero cuadrado tras validación BD)
    idx_gestion = col_index("Gestión a Realizar") or col_index("Gestion a Realizar")
    if idx_gestion is not None and (idx_diferencia == idx_gestion or (idx_naturaleza is not None and idx_naturaleza == idx_gestion)):
        logger.warning("Columna Diferencia/Naturaleza coincide con Gestión a Realizar; no se escriben fórmulas para no sobrescribir.")
        return

    for row in range(2, sheet.max_row + 1):
        parte_remanente = f"- {letra_remanente}{row}" if letra_remanente else ""
        formula = f"={letra_saldo}{row}-({letra_efectivo}{row}+{letra_dispensado}{row}-{letra_recibido}{row}){parte_remanente}"
        sheet.cell(row=row, column=idx_diferencia).value = formula
        # Naturaleza: si Diferencia=0 "Cuadrado", si >0 "Faltante", sino "Sobrante"
        if idx_naturaleza is not None:
            formula_nat = f'=IF({letra_diferencia}{row}=0,"Cuadrado",IF({letra_diferencia}{row}>0,"Faltante","Sobrante"))'
            sheet.cell(row=row, column=idx_naturaleza).value = formula_nat

    wb.save(ruta)
    logger.info("Fórmula Diferencia escrita en columna %s (filas 2 a %d)", letra_diferencia, sheet.max_row)
    if idx_naturaleza is not None:
        logger.info("Fórmula Naturaleza escrita (Cuadrado/Faltante/Sobrante según Diferencia).")


def _escribir_formulas_remanente_en_excel(
    ruta: Path,
    filas_formula: list,
    nombre_hoja: str = "DETALLE MF",
) -> None:
    """
    Escribe en la columna 'Remanente /Provisión /Ajustes' las fórmulas indicadas.
    filas_formula: lista de (número_fila_excel, fórmula) donde fórmula es string tipo "=valor1-valor2+valor3".
    """
    if not filas_formula:
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl no disponible para escribir fórmulas Remanente")
        return
    wb = load_workbook(ruta)
    if nombre_hoja not in wb.sheetnames:
        nombre_hoja = wb.sheetnames[0]
    sheet = wb[nombre_hoja]
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]

    def col_index(nombre: str) -> Optional[int]:
        n = (nombre or "").strip().lower()
        for i, h in enumerate(headers):
            if (str(h or "").strip().lower() == n):
                return i + 1
        # Alias sin espacios
        if "remanente" in n and "provision" in n:
            for i, h in enumerate(headers):
                if h and "remanente" in (str(h) or "").lower() and "provision" in (str(h) or "").lower():
                    return i + 1
        return None

    idx_remanente = col_index("Remanente /Provisión /Ajustes") or col_index("Remanente/Provisión/Ajustes")
    if idx_remanente is None:
        logger.warning("No se encontró columna Remanente/Provisión/Ajustes para escribir fórmulas.")
        return
    for excel_row, formula in filas_formula:
        if formula and str(formula).strip().startswith("="):
            sheet.cell(row=excel_row, column=idx_remanente).value = formula
    wb.save(ruta)
    logger.info("Fórmulas Remanente escritas en %d celda(s).", len(filas_formula))


def _rellenar_marca_desde_lista_mf(ruta: Path) -> None:
    """
    Rellena la columna Marca en la hoja DETALLE MF con el valor de la hoja Lista MF:
    para cada fila en DETALLE MF, busca el Cajero en Lista MF (columna Cajero) y copia
    el valor de la columna Marca de Lista MF en la columna Marca de DETALLE MF.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl no disponible para rellenar Marca desde Lista MF")
        return
    wb = load_workbook(ruta)
    if "Lista MF" not in wb.sheetnames:
        logger.warning("Hoja 'Lista MF' no encontrada; no se rellena columna Marca.")
        return
    if "DETALLE MF" not in wb.sheetnames:
        logger.warning("Hoja 'DETALLE MF' no encontrada; no se rellena columna Marca.")
        return

    sheet_lista = wb["Lista MF"]
    headers_lista = [sheet_lista.cell(row=1, column=c).value for c in range(1, sheet_lista.max_column + 1)]

    def col_index_lista(nombre: str) -> Optional[int]:
        n = (nombre or "").strip().lower()
        for i, h in enumerate(headers_lista):
            if (str(h or "").strip().lower() == n):
                return i + 1
        return None

    idx_cajero_lista = col_index_lista("Cajero")
    idx_marca_lista = col_index_lista("Marca")
    if not idx_cajero_lista or not idx_marca_lista:
        logger.warning("En Lista MF no se encontraron columnas 'Cajero' y/o 'Marca'.")
        return

    # Construir mapeo Cajero -> Marca desde Lista MF
    cajero_a_marca: Dict[str, Any] = {}
    for row in range(2, sheet_lista.max_row + 1):
        cajero = sheet_lista.cell(row=row, column=idx_cajero_lista).value
        marca = sheet_lista.cell(row=row, column=idx_marca_lista).value
        if cajero is not None and str(cajero).strip() != "":
            clave = str(cajero).strip()
            cajero_a_marca[clave] = marca

    sheet_detalle = wb["DETALLE MF"]
    headers_detalle = [sheet_detalle.cell(row=1, column=c).value for c in range(1, sheet_detalle.max_column + 1)]

    def col_index_detalle(nombre: str) -> Optional[int]:
        n = (nombre or "").strip().lower()
        for i, h in enumerate(headers_detalle):
            if (str(h or "").strip().lower() == n):
                return i + 1
        return None

    idx_cajero_detalle = col_index_detalle("Cajero")
    idx_marca_detalle = col_index_detalle("Marca")
    if not idx_cajero_detalle or not idx_marca_detalle:
        logger.warning("En DETALLE MF no se encontraron columnas 'Cajero' y/o 'Marca'.")
        return

    rellenadas = 0
    for row in range(2, sheet_detalle.max_row + 1):
        cajero = sheet_detalle.cell(row=row, column=idx_cajero_detalle).value
        if cajero is not None and str(cajero).strip() != "":
            clave = str(cajero).strip()
            if clave in cajero_a_marca:
                sheet_detalle.cell(row=row, column=idx_marca_detalle).value = cajero_a_marca[clave]
                rellenadas += 1

    wb.save(ruta)
    logger.info("Columna Marca en DETALLE MF rellenada desde Lista MF (%d filas con coincidencia).", rellenadas)


def _construir_columnas_mapeo(df: pd.DataFrame) -> Dict[str, str]:
    """Construye mapeo columna real en df -> columna en ARQUEOS MF usando MAPEO_A_ARQUEOS_MF y ALIASES_ORIGEN."""
    columnas_origen = {}
    for nombre_posible, col_arqueos in MAPEO_A_ARQUEOS_MF.items():
        variantes = [nombre_posible, nombre_posible.replace(" ", "_"), nombre_posible.replace("_", " ")]
        if nombre_posible in ALIASES_ORIGEN:
            variantes = variantes + list(ALIASES_ORIGEN[nombre_posible])
        col = _encontrar_columna(df, variantes)
        if col is not None:
            columnas_origen[col] = col_arqueos
    return columnas_origen


def _pegar_filas_a_arqueos_mf(
    df_nuevas: pd.DataFrame,
    ruta_arqueos_mf: Optional[Union[str, Path]] = None,
    lector=None,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
    hoja_arqueos_mf: Union[str, int] = 0,
    etiqueta_log: str = "filas",
) -> pd.DataFrame:
    """Carga ARQUEOS MF, concatena df_nuevas al final, guarda y devuelve el resultado."""
    if ruta_arqueos_mf is not None:
        ruta_arqueos_mf = Path(ruta_arqueos_mf)
    if ruta_arqueos_mf and ruta_arqueos_mf.exists():
        df_mf = pd.read_excel(ruta_arqueos_mf, sheet_name=hoja_arqueos_mf, engine="openpyxl")
        ruta_save = ruta_arqueos_mf
    elif lector:
        df_mf = lector.leer_arqueos_mf(mes=mes, anio=anio, hoja=hoja_arqueos_mf)
        ruta_save = lector._ruta_arqueos_mf(mes=mes, anio=anio)
    else:
        raise ValueError("Indique ruta_arqueos_mf o lector con mes/anio")

    # Quitar filas vacías al final (fórmulas sin datos) para que al pegar no queden espacios en blanco antes de los nuevos registros
    def _col_similar(cols: list, nombre: str) -> Optional[str]:
        n = (nombre or "").strip().lower()
        for c in cols:
            if (str(c or "").strip().lower() == n):
                return c
        return None

    cols_mf = list(df_mf.columns)
    col_cajero_mf = _col_similar(cols_mf, "Cajero")
    col_fecha_mf = _col_similar(cols_mf, "Fecha Arqueo")
    col_saldo_mf = _col_similar(cols_mf, "Saldo contable") or _col_similar(cols_mf, "Saldo Contable")
    if df_mf.shape[0] > 0 and (col_cajero_mf or col_fecha_mf or col_saldo_mf):
        def _tiene_valor_mf(serie: pd.Series) -> pd.Series:
            if serie is None or len(serie) == 0:
                return pd.Series(False, index=df_mf.index)
            return serie.notna() & (serie.astype(str).str.strip() != "") & (serie.astype(str).str.strip().str.lower() != "nan")
        tiene_dato = pd.Series(False, index=df_mf.index)
        for col in (col_cajero_mf, col_fecha_mf, col_saldo_mf):
            if col is not None and col in df_mf.columns:
                tiene_dato = tiene_dato | _tiene_valor_mf(df_mf[col])
        posiciones_con_dato = (tiene_dato.values).nonzero()[0]
        ultima_pos = int(posiciones_con_dato[-1]) if len(posiciones_con_dato) > 0 else -1
        if ultima_pos >= 0 and ultima_pos < len(df_mf) - 1:
            n_eliminadas = len(df_mf) - (ultima_pos + 1)
            df_mf = df_mf.iloc[: ultima_pos + 1].copy()
            logger.info("ARQUEOS MF recortado a %d filas (eliminadas %d filas vacías al final).", len(df_mf), n_eliminadas)

    # Alinear nombres de df_nuevas a los del destino (ej. "Saldo Contable" -> "Saldo contable")
    # para no crear columnas nuevas y pegar en las existentes
    def _nombre_en_destino(destino_columns: list, nombre: str) -> Optional[str]:
        n = (nombre or "").strip().lower()
        for c in destino_columns:
            if (c or "").strip().lower() == n:
                return c
        return None

    renombrar = {}
    for col in df_nuevas.columns:
        nombre_real = _nombre_en_destino(list(df_mf.columns), col)
        if nombre_real is not None and nombre_real != col:
            renombrar[col] = nombre_real
    if renombrar:
        df_nuevas = df_nuevas.rename(columns=renombrar)

    for c in df_mf.columns:
        if c not in df_nuevas.columns:
            df_nuevas[c] = pd.NA
    df_nuevas = df_nuevas.reindex(columns=df_mf.columns, fill_value=pd.NA)

    # No pegar filas que quedaron sin datos en columnas clave (evita filas vacías por mapeo incorrecto del consolidado)
    def _col_similar(cols: list, nombre: str) -> Optional[str]:
        n = (nombre or "").strip().lower()
        for c in cols:
            if (str(c or "").strip().lower() == n):
                return c
        return None
    col_cajero = _col_similar(list(df_nuevas.columns), "Cajero")
    col_fecha = _col_similar(list(df_nuevas.columns), "Fecha Arqueo")
    col_saldo = _col_similar(list(df_nuevas.columns), "Saldo contable") or _col_similar(list(df_nuevas.columns), "Saldo Contable")
    if col_cajero is not None or col_fecha is not None or col_saldo is not None:
        def _tiene_valor(serie: pd.Series) -> pd.Series:
            if serie is None or len(serie) == 0:
                return pd.Series(False, index=df_nuevas.index)
            return serie.notna() & (serie.astype(str).str.strip() != "") & (serie.astype(str).str.strip().str.lower() != "nan")
        tiene_algo = pd.Series(False, index=df_nuevas.index)
        for col in (col_cajero, col_fecha, col_saldo):
            if col is not None:
                tiene_algo = tiene_algo | _tiene_valor(df_nuevas[col])
        n_omitir = (~tiene_algo).sum()
        df_nuevas = df_nuevas[tiene_algo].copy()
        if n_omitir > 0:
            logger.warning("Se omitieron %d fila(s) sin datos en Cajero/Fecha Arqueo/Saldo contable (mapeo incompleto).", int(n_omitir))

    df_resultado = pd.concat([df_mf, df_nuevas], ignore_index=True)
    logger.info("Pegadas %d %s al final de ARQUEOS MF. Total filas: %d", len(df_nuevas), etiqueta_log, len(df_resultado))

    # El texto "Cajero cuadrado..." se graba solo después de la validación en BD (movimientos 770500/810291 y Remanente)
    if lector:
        lector.guardar_arqueos_mf(df_resultado, mes=mes, anio=anio, hoja=hoja_arqueos_mf)
    else:
        if ruta_save.exists():
            with pd.ExcelFile(ruta_save, engine="openpyxl") as xl:
                nombres_hojas = xl.sheet_names
            reemplazar = (lambda i, n: i == hoja_arqueos_mf) if isinstance(hoja_arqueos_mf, int) else (lambda i, n: n == hoja_arqueos_mf)
            dict_hojas = {}
            for i, nombre in enumerate(nombres_hojas):
                if reemplazar(i, nombre):
                    dict_hojas[nombre] = df_resultado
                else:
                    dict_hojas[nombre] = pd.read_excel(ruta_save, sheet_name=nombre, engine="openpyxl")
            with pd.ExcelWriter(ruta_save, engine="openpyxl") as writer:
                for nombre, frame in dict_hojas.items():
                    frame.to_excel(writer, sheet_name=nombre, index=False)
        else:
            df_resultado.to_excel(ruta_save, index=False, engine="openpyxl")

    # Escribir fórmula Diferencia en toda la hoja
    _escribir_formula_diferencia_en_excel(ruta_save)
    # Rellenar columna Marca en DETALLE MF desde Lista MF (Cajero -> Marca)
    _rellenar_marca_desde_lista_mf(ruta_save)

    return df_resultado


def pegar_gestion_a_arqueos_mf(
    df_gestion: pd.DataFrame,
    ruta_arqueos_mf: Optional[str] = None,
    lector=None,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
    hoja_arqueos_mf: Union[str, int] = 0,
) -> pd.DataFrame:
    """
    Filtra el archivo de gestión por tipo_registro = ARQUEO, mapea columnas y pega
    las filas al final del archivo ARQUEOS MF.

    Args:
        df_gestion: DataFrame del archivo de gestión (erestrad).
        ruta_arqueos_mf: Ruta al archivo ARQUEOS MF. Si None, se usa lector + mes/anio.
        lector: LectorInsumos (necesario si ruta_arqueos_mf es None).
        mes: Mes 1-12 para el archivo ARQUEOS MF.
        anio: Año para el archivo ARQUEOS MF.
        hoja_arqueos_mf: Hoja a editar en ARQUEOS MF (0 o nombre).

    Returns:
        DataFrame del archivo ARQUEOS MF después de pegar (con las filas nuevas al final).
    """
    ruta_arqueos_mf = Path(ruta_arqueos_mf) if ruta_arqueos_mf else None
    if "tipo_registro" not in df_gestion.columns:
        raise ValueError("El archivo de gestión debe tener la columna 'tipo_registro'")
    df_arqueo = df_gestion[df_gestion["tipo_registro"].astype(str).str.strip().str.upper() == "ARQUEO"].copy()
    if df_arqueo.empty:
        logger.warning("No hay registros con tipo_registro = ARQUEO en el archivo de gestión.")
        if ruta_arqueos_mf and ruta_arqueos_mf.exists():
            return pd.read_excel(ruta_arqueos_mf, sheet_name=hoja_arqueos_mf, engine="openpyxl")
        if lector:
            return lector.leer_arqueos_mf(mes=mes, anio=anio, hoja=hoja_arqueos_mf)
        raise ValueError("No hay datos para pegar y no se pudo leer ARQUEOS MF")

    columnas_origen = _construir_columnas_mapeo(df_gestion)
    filas_nuevas = [_mapear_fila_a_arqueos_mf(row, columnas_origen) for _, row in df_arqueo.iterrows()]
    if not filas_nuevas:
        logger.warning("No se mapeó ninguna columna; revisar nombres en el archivo de gestión.")
        if ruta_arqueos_mf and ruta_arqueos_mf.exists():
            return pd.read_excel(ruta_arqueos_mf, sheet_name=hoja_arqueos_mf, engine="openpyxl")
        return lector.leer_arqueos_mf(mes=mes, anio=anio, hoja=hoja_arqueos_mf)

    df_nuevas = pd.DataFrame(filas_nuevas)
    return _pegar_filas_a_arqueos_mf(
        df_nuevas,
        ruta_arqueos_mf=ruta_arqueos_mf,
        lector=lector,
        mes=mes,
        anio=anio,
        hoja_arqueos_mf=hoja_arqueos_mf,
        etiqueta_log="filas de gestión (ARQUEO)",
    )


def _parsear_fecha_dd_mm_yyyy(fecha: str):
    """Convierte DD_MM_YYYY o DD-MM-YYYY a datetime.date."""
    from datetime import datetime
    s = fecha.strip().replace("-", "_")
    partes = s.split("_")
    if len(partes) != 3:
        raise ValueError(f"Fecha debe ser DD_MM_YYYY: {fecha}")
    d, m, a = int(partes[0]), int(partes[1]), int(partes[2])
    return datetime(a, m, d).date()


def pegar_consolidado_a_arqueos_mf(
    df_consolidado: pd.DataFrame,
    ruta_arqueos_mf: Optional[Union[str, Path]] = None,
    lector=None,
    mes: Optional[int] = None,
    anio: Optional[int] = None,
    hoja_arqueos_mf: Union[str, int] = 0,
    fecha_proceso: Optional[str] = None,
) -> pd.DataFrame:
    """
    Filtra el consolidado por tipo_cajero que contenga "multifuncional" y tipo_registro = "ARQUEO",
    mapea con la misma asociación a ARQUEOS MF y pega las filas al final.
    Si "Fecha descarga arqueo" queda vacía (no viene en consolidado), se rellena con fecha_proceso.

    Args:
        df_consolidado: DataFrame del archivo consolidado_cajeros_cuadrados.
        ruta_arqueos_mf: Ruta al archivo ARQUEOS MF. Si None, se usa lector + mes/anio.
        lector: LectorInsumos (necesario si ruta_arqueos_mf es None).
        mes: Mes 1-12 para el archivo ARQUEOS MF.
        anio: Año para el archivo ARQUEOS MF.
        hoja_arqueos_mf: Hoja a editar en ARQUEOS MF (0 o nombre).
        fecha_proceso: Fecha DD_MM_YYYY (ej. del archivo de gestión) para rellenar "Fecha descarga arqueo" si viene vacía.

    Returns:
        DataFrame del archivo ARQUEOS MF después de pegar.
    """
    if ruta_arqueos_mf is not None:
        ruta_arqueos_mf = Path(ruta_arqueos_mf)
    if "tipo_cajero" not in df_consolidado.columns:
        raise ValueError("El consolidado debe tener la columna 'tipo_cajero'")
    tipo_str = df_consolidado["tipo_cajero"].astype(str).str.strip().str.upper()
    mask_mf = tipo_str.str.contains("MULTIFUNCIONAL", na=False)
    if "tipo_registro" in df_consolidado.columns:
        mask_arqueo = df_consolidado["tipo_registro"].astype(str).str.strip().str.upper() == "ARQUEO"
        df_filtrado = df_consolidado[mask_mf & mask_arqueo].copy()
    else:
        logger.warning("El consolidado no tiene columna 'tipo_registro'; se filtran solo por tipo_cajero multifuncional.")
        df_filtrado = df_consolidado[mask_mf].copy()
    if df_filtrado.empty:
        logger.warning("No hay registros multifuncionales con tipo_registro = ARQUEO en el consolidado.")
        if ruta_arqueos_mf and ruta_arqueos_mf.exists():
            return pd.read_excel(ruta_arqueos_mf, sheet_name=hoja_arqueos_mf, engine="openpyxl")
        if lector:
            return lector.leer_arqueos_mf(mes=mes, anio=anio, hoja=hoja_arqueos_mf)
        raise ValueError("No hay datos para pegar y no se pudo leer ARQUEOS MF")

    columnas_origen = _construir_columnas_mapeo(df_consolidado)
    filas_nuevas = [_mapear_fila_a_arqueos_mf(row, columnas_origen) for _, row in df_filtrado.iterrows()]
    if not filas_nuevas:
        logger.warning("No se mapeó ninguna columna; revisar nombres en el consolidado.")
        if ruta_arqueos_mf and ruta_arqueos_mf.exists():
            return pd.read_excel(ruta_arqueos_mf, sheet_name=hoja_arqueos_mf, engine="openpyxl")
        return lector.leer_arqueos_mf(mes=mes, anio=anio, hoja=hoja_arqueos_mf)

    df_nuevas = pd.DataFrame(filas_nuevas)

    # Fecha descarga arqueo = misma que gestión (fecha_proceso). El consolidado no la trae, se pone para todos.
    if fecha_proceso:
        try:
            fecha_valor = _parsear_fecha_dd_mm_yyyy(fecha_proceso)
            col_fecha = next(
                (c for c in df_nuevas.columns if str(c).strip().lower() == "fecha descarga arqueo"),
                None,
            )
            if col_fecha is not None:
                mask_vacia = df_nuevas[col_fecha].isna() | (df_nuevas[col_fecha].astype(str).str.strip().isin(["", "nan"]))
                df_nuevas.loc[mask_vacia, col_fecha] = fecha_valor
            else:
                # La columna no existe en consolidado; crearla y llenarla con la fecha de proceso
                df_nuevas["Fecha descarga arqueo"] = fecha_valor
        except ValueError:
            pass

    return _pegar_filas_a_arqueos_mf(
        df_nuevas,
        ruta_arqueos_mf=ruta_arqueos_mf,
        lector=lector,
        mes=mes,
        anio=anio,
        hoja_arqueos_mf=hoja_arqueos_mf,
        etiqueta_log="filas de consolidado (multifuncional + ARQUEO)",
    )
